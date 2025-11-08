from __future__ import annotations

import io
import json
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import Any, Dict, Optional

from fastapi import Depends, FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session

from . import crud
from .config import get_settings
from .database import Base, engine, get_session
from .schemas import (
    BreakEvenResponse,
    ExcelImportResponse,
    ExcelImportWarning,
    PriceSimulationRequest,
    PriceSimulationResponse,
)
from .utils import generate_price_patterns, round_jpy, round_rate

settings = get_settings()

app = FastAPI(title="Pricing Decision Support System")
app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.allowed_cors_origins,
    allow_methods=["*"],
    allow_headers=["*"],
    allow_credentials=True,
)

Base.metadata.create_all(bind=engine)

ERROR_INVALID_PARAM = {
    "error": {
        "code": "INVALID_PARAM",
        "message": "入力値が不正です",
    }
}

DEFAULT_COLUMN_MAPPING = {
    "product_code": "C",
    "product_name": "D",
    "category": "E",
    "unit_cost_per_kg": "F",
    "unit_price_per_kg": "G",
    "target_margin_rate": "H",
    "min_margin_rate": "I",
}


@app.post("/api/price-simulations/calculate", response_model=PriceSimulationResponse)
def calculate_price_simulation(
    payload: PriceSimulationRequest,
) -> PriceSimulationResponse:
    try:
        unit_cost = Decimal(str(payload.unit_cost_per_kg))
        target_margin_rate = Decimal(str(payload.target_margin_rate))
        quantity = (
            Decimal(str(payload.quantity_kg))
            if payload.quantity_kg is not None
            else None
        )
    except (InvalidOperation, TypeError) as exc:
        raise HTTPException(status_code=400, detail=ERROR_INVALID_PARAM) from exc

    if unit_cost <= Decimal("0"):
        raise HTTPException(
            status_code=400,
            detail={
                "error": {
                    "code": "INVALID_PARAM",
                    "message": "unit_cost_per_kg must be greater than 0",
                }
            },
        )
    if target_margin_rate < Decimal("0") or target_margin_rate >= Decimal("1"):
        raise HTTPException(
            status_code=400,
            detail={
                "error": {
                    "code": "INVALID_PARAM",
                    "message": "target_margin_rate must be between 0.0 and 0.9",
                }
            },
        )
    if quantity is not None and quantity < Decimal("0"):
        raise HTTPException(
            status_code=400,
            detail={
                "error": {
                    "code": "INVALID_PARAM",
                    "message": "quantity_kg must be greater than or equal to 0",
                }
            },
        )

    one = Decimal("1")
    recommended_price = unit_cost / (one - target_margin_rate)
    gross_profit_per_kg = recommended_price - unit_cost

    price_patterns = [
        {
            "margin_rate": float(round_rate(margin_rate)),
            "price_per_kg": price_per_kg,
            "profit_per_kg": profit_per_kg,
        }
        for margin_rate, price_per_kg, profit_per_kg in generate_price_patterns(unit_cost)
    ]

    gross_profit_total: Optional[int] = None
    if quantity is not None:
        gross_profit_total = round_jpy(gross_profit_per_kg * quantity)

    recommended_price_rounded = round_jpy(recommended_price)
    gross_profit_per_kg_rounded = round_jpy(gross_profit_per_kg)

    return PriceSimulationResponse(
        recommended_price_per_kg=recommended_price_rounded,
        gross_profit_per_kg=gross_profit_per_kg_rounded,
        gross_profit_total=gross_profit_total,
        margin_rate=float(round_rate(target_margin_rate)),
        price_patterns=price_patterns,
        guard={
            "min_allowed_price_per_kg": recommended_price_rounded,
            "is_below_min": False,
            "warning_message": None,
        },
    )


@app.get("/api/break-even/current", response_model=BreakEvenResponse)
def get_break_even(
    year_month: str,
    session: Session = Depends(get_session),
) -> BreakEvenResponse:
    try:
        year = int(year_month.split("-")[0])
        month = int(year_month.split("-")[1])
        month_start = date(year, month, 1)
    except Exception as exc:  # pragma: no cover - validation
        raise HTTPException(
            status_code=400,
            detail={"error": {"code": "INVALID_PARAM", "message": "Invalid year_month"}},
        ) from exc

    # Determine next month
    if month == 12:
        month_end = date(year + 1, 1, 1)
    else:
        month_end = date(year, month + 1, 1)

    fixed_cost_total = crud.get_fixed_cost_total(session, month_start)
    revenue, variable_cost = crud.get_sales_summary(session, month_start, month_end)

    variable_cost_rate = (
        round_rate(variable_cost / revenue) if revenue > 0 else Decimal("0")
    )
    gross_margin_rate_raw = (
        Decimal("1") - (variable_cost / revenue if revenue > 0 else Decimal("0"))
    )
    gross_margin_rate = (
        round_rate(gross_margin_rate_raw) if gross_margin_rate_raw > 0 else Decimal("0")
    )

    if gross_margin_rate_raw <= 0:
        break_even_revenue = None
        achievement_rate = Decimal("0")
        delta_revenue = -fixed_cost_total
    else:
        break_even_value = fixed_cost_total / gross_margin_rate_raw
        break_even_revenue = break_even_value
        achievement_rate = (
            revenue / break_even_value if break_even_value > 0 else Decimal("0")
        )
        delta_revenue = revenue - break_even_value

    if achievement_rate >= Decimal("1"):
        status = "safe"
    elif achievement_rate >= Decimal("0.8"):
        status = "warning"
    else:
        status = "danger"

    return BreakEvenResponse(
        year_month=year_month,
        fixed_costs=round_jpy(fixed_cost_total),
        current_revenue=round_jpy(revenue),
        variable_cost_rate=float(variable_cost_rate),
        gross_margin_rate=float(gross_margin_rate),
        break_even_revenue=round_jpy(break_even_revenue)
        if break_even_revenue is not None
        else 0,
        achievement_rate=float(round_rate(achievement_rate))
        if achievement_rate > 0
        else 0.0,
        delta_revenue=round_jpy(delta_revenue),
        status=status,
    )


def _parse_column_mapping(column_mapping: Optional[str]) -> Dict[str, str]:
    if not column_mapping:
        return DEFAULT_COLUMN_MAPPING
    try:
        data = json.loads(column_mapping)
        if not isinstance(data, dict):
            raise ValueError
        normalized = {key: value.upper() for key, value in data.items()}
        return {**DEFAULT_COLUMN_MAPPING, **normalized}
    except ValueError as exc:  # pragma: no cover - invalid mapping path
        raise HTTPException(
            status_code=400,
            detail={
                "error": {
                    "code": "INVALID_PARAM",
                    "message": "column_mapping must be a JSON object",
                }
            },
        ) from exc


def _cell_value(sheet, column_letter: str, row_index: int) -> Any:
    cell = sheet[f"{column_letter}{row_index}"]
    return cell.value


def _normalize_rate(value: Any) -> Optional[Decimal]:
    if value is None:
        return None
    try:
        numeric = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None
    if numeric > Decimal("1"):
        numeric = numeric / Decimal("100")
    if numeric < Decimal("0"):
        return None
    return round_rate(numeric)


def _normalize_currency(value: Any) -> Optional[Decimal]:
    if value is None:
        return None
    try:
        numeric = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None
    return (numeric * Decimal("1000")).quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)


@app.post("/api/import/excel", response_model=ExcelImportResponse)
async def import_excel(
    file: UploadFile = File(...),
    column_mapping: Optional[str] = Form(default=None),
    session: Session = Depends(get_session),
) -> ExcelImportResponse:
    from openpyxl import load_workbook

    content = await file.read()
    try:
        workbook = load_workbook(filename=io.BytesIO(content), data_only=True)
    except Exception as exc:  # pragma: no cover - invalid file
        raise HTTPException(
            status_code=400,
            detail={"error": {"code": "INVALID_FILE", "message": "Failed to read Excel file"}},
        ) from exc

    sheet = workbook.active
    mapping = _parse_column_mapping(column_mapping)

    imported = 0
    skipped = 0
    warnings: list[ExcelImportWarning] = []

    for row in range(2, sheet.max_row + 1):
        product_code = _cell_value(sheet, mapping["product_code"], row)
        product_name = _cell_value(sheet, mapping["product_name"], row)

        if not product_code or not product_name:
            skipped += 1
            warnings.append(
                ExcelImportWarning(
                    row=row,
                    field="product_code",
                    reason="missing product_code or product_name",
                )
            )
            continue

        unit_cost = _normalize_currency(_cell_value(sheet, mapping["unit_cost_per_kg"], row))
        unit_price = _normalize_currency(_cell_value(sheet, mapping["unit_price_per_kg"], row))

        if unit_cost is None or unit_price is None:
            skipped += 1
            warnings.append(
                ExcelImportWarning(
                    row=row,
                    field="unit_cost_per_kg",
                    reason="non-numeric",
                )
            )
            continue

        target_margin_rate = _normalize_rate(
            _cell_value(sheet, mapping["target_margin_rate"], row)
        )
        min_margin_rate = _normalize_rate(
            _cell_value(sheet, mapping.get("min_margin_rate", ""), row)
            if mapping.get("min_margin_rate")
            else None
        )
        category = _cell_value(sheet, mapping.get("category", ""), row)

        product_data = {
            "product_code": str(product_code).strip(),
            "product_name": str(product_name).strip(),
            "category": str(category).strip() if category else None,
            "unit_cost_per_kg": unit_cost,
            "unit_price_per_kg": unit_price,
            "target_margin_rate": target_margin_rate,
            "min_margin_rate": min_margin_rate,
            "unit": "JPY/kg",
        }

        crud.upsert_product(session, product_data)
        imported += 1

    session.commit()

    return ExcelImportResponse(
        imported=imported,
        skipped=skipped,
        warnings=warnings,
    )
